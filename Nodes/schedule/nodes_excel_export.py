from bpy.types import Node
from bpy.props import StringProperty, CollectionProperty, IntProperty

from .tree import MaStroScheduleTreeNode, resolve_origin_node, resolve_through_reroutes, resolve_named_origin
from .execution import get_node_table
from .properties import MaStro_schedule_export_sheet_item


def _link_key(from_node, output_index, link_position):
    """Same stable string identity as Join Tables/Join Sheets' own
    _link_key - see nodes_table_join.py's own docstring, including why
    link_position (this link's own index within socket.links) is part
    of the key now too."""
    return f"{from_node.name}::{output_index}::{link_position}"


def _label_text(sheet):
    """Same purely-cosmetic first-cell-text label as Join Sheets'
    own _label_text (nodes_sheet_place.py) - see that function's own
    docstring."""
    columns = sheet.get("columns", []) if sheet else []
    if not columns:
        return ""
    cells = columns[0].get("cells", [])
    return cells[0].get("text", "") if cells else ""


def _origin_label(origin_node, sheet):
    """origin_node's own table_or_sheet_name (Join Sheets/Join Tables'
    own shared optional name property) takes priority over
    _label_text's first-cell-text fallback - the user's own explicit
    fix: a Join Sheets' combined result showed "(empty)" (whichever
    linked Sheet happened to be first had no text of its own in its
    first cell), meaningless for a node whose whole job is combining
    several Sheets, not being any one of them.

    If origin_node itself has no name, resolve_named_origin walks
    FURTHER upstream looking for one - confirmed live as a real gap
    otherwise: Join Tables -> Table to Sheet -> Export Excel never saw
    Join Tables' own table_or_sheet_name at all, since Table to Sheet
    (origin_node here, a type-change boundary resolve_origin_node
    correctly stops at for IDENTITY) has no name of its own to check.
    See resolve_named_origin's own docstring in tree.py for why this
    is a separate walk, only for the displayed label, never for
    sheet_items' own link_key."""
    custom_name = getattr(origin_node, "table_or_sheet_name", "") or resolve_named_origin(origin_node)
    if custom_name:
        return custom_name
    return _label_text(sheet)


# Writes every linked Sheet into one .xlsx workbook, one Excel sheet
# per linked Sheet - the user's own framing: "ogni viewer potrebbe
# corrispondere a una sheet diversa di excel, quindi il nodo deve
# prevedere input multipli, uno per sheet". Same real Blender
# multi-input socket as Join Tables/Join Sheets, but with its own
# dedicated PropertyGroup/UIList/Move operator
# (MaStro_schedule_export_sheet_item/MASTRO_UL_schedule_export_sheets/
# MASTRO_OT_Schedule_Export_Sheets_Move) rather than reusing Join
# Tables' own table_items - the user's own explicit call: those two
# nodes have no use for sheet_name/mode/start_cell, reusing one
# PropertyGroup for both would leave dead properties on unrelated
# nodes.
#
# No editable name field on this list anymore (see _origin_label/
# export_sheets' own comments below) - naming a Sheet happens upstream,
# once, at Table to Sheet/Join Tables/Join Sheets (all sharing one
# table_or_sheet_name property, see resolve_named_origin's own
# docstring in tree.py). The user's own explicit simplification, once
# every one of those nodes could carry a name of its own: a second
# rename field here, downstream of all of them, was redundant.
#
# No formulas - the user's own explicit call - and no gamma/color
# management on the RGB->hex conversion - direct linear mapping (see
# excel_export_shared.py's own docstring for both).
#
# Export only ever happens manually, via the Export button
# (MASTRO_OT_Schedule_Excel_Export below) - reads whatever this node's
# own evaluate() last cached, never through Blender's own export
# operator machinery (which would show a "file already exists"
# confirmation popup, not the silent-on-success/footer-INFO-on-success
# behavior the button has instead). An auto_export option (writing
# silently from tree.py's own polling timer on every graph change) was
# tried and removed - the user's own explicit reversal, after a past
# bad experience with auto-export silently overwriting Excel data that
# shouldn't have been touched; manual-only, with the user always in
# control of when a write actually happens, was the safer call. Writes
# atomically regardless (excel_export_shared.save_workbook_atomically)
# - a temp file in the same directory, then os.replace() into place -
# so a crash mid-save never leaves a half-written .xlsx at the real
# path.
class MaStroScheduleExcelExportNode(MaStroScheduleTreeNode, Node):
    """Export every linked Sheet as one sheet each in a single .xlsx
    workbook"""
    bl_idname = 'MaStroScheduleExcelExport'
    bl_label = 'Export to Excel'
    # Wider than the default (~140px) - the user's own explicit call,
    # this node's own filepath field/sheet_items list both need more
    # room than that to stay readable.
    bl_width_default = 300

    # No default path - the user's own call: "//export.xlsx" (a path
    # relative to the .blend file) doesn't resolve to anything real
    # until the file is actually saved somewhere, which made the field
    # show Blender's own native "this file doesn't exist" red
    # background right away on a brand new node - left blank instead,
    # so the only time that red background appears is once the user
    # has actually typed something that doesn't resolve, not as a
    # false alarm on every freshly added node.
    filepath: StringProperty(name="Path", subtype='FILE_PATH', default="")
    sheet_items: CollectionProperty(type=MaStro_schedule_export_sheet_item)
    active_sheet_index: IntProperty()

    def init(self, context):
        # use_multi_input is a constructor-only argument - see Join
        # Tables' own init() (nodes_table_join.py) for the confirmed
        # RNA source detail.
        self.inputs.new('MaStroScheduleSheetSocketType', "Sheet", use_multi_input=True)

    def _sync_sheet_items(self):
        """Same sync logic as Join Tables/Join Sheets' own
        _sync_table_items - see nodes_table_join.py's own docstring,
        and resolve_origin_node's own docstring in tree.py, for the
        full story. Must NOT be called from draw_buttons (same
        "Writing to ID classes in this context is not allowed"
        restriction) - runs from tree.py's own polling timer instead."""
        socket = self.inputs["Sheet"]
        current_keys = []
        labels_by_key = {}
        for link_position, link in enumerate(socket.links):
            if link.is_muted:
                continue
            origin_node, origin_socket = resolve_origin_node(link)
            if origin_node is None:
                continue
            try:
                output_index = list(origin_node.outputs).index(origin_socket)
            except ValueError:
                continue
            key = _link_key(origin_node, output_index, link_position)
            current_keys.append(key)
            sheet = get_node_table(self.id_data.name, origin_node.name)
            labels_by_key[key] = _origin_label(origin_node, sheet[output_index] if sheet else None)

        existing_keys = [item.link_key for item in self.sheet_items]
        for index in reversed(range(len(self.sheet_items))):
            if existing_keys[index] not in current_keys:
                self.sheet_items.remove(index)
        tracked_keys = {item.link_key for item in self.sheet_items}
        for key in current_keys:
            if key not in tracked_keys:
                item = self.sheet_items.add()
                item.link_key = key
                tracked_keys.add(key)
        for item in self.sheet_items:
            item.label = labels_by_key.get(item.link_key, "")

    def _has_valid_filepath(self):
        """False for an empty path, a path ending in a slash (the
        file browser's own folder-only selection, confirmed live as
        the exact case the user typed in without ever noticing
        nothing was wrong - the Export button stayed enabled and
        appeared to work), or a path that already exists as a real
        directory on disk (e.g. typed without a trailing slash, but
        still naming an existing folder, not a file inside it) -
        os.path.basename() alone doesn't catch that last case (a bare
        folder name with no slash still returns a non-empty
        basename)."""
        import os
        import bpy
        path = self.filepath
        if not path or path.endswith(("/", "\\")):
            return False
        if not os.path.basename(path):
            return False
        return not os.path.isdir(bpy.path.abspath(path))

    def draw_buttons(self, context, layout):
        layout.prop(self, "filepath")
        export_row = layout.row()
        # Disabled rather than letting the operator run with an empty/
        # folder-only path - the user's own explicit call: nothing
        # useful to export TO without a real file name, so the button
        # itself should make that obvious rather than only failing (or
        # worse, silently writing somewhere unintended) once clicked.
        export_row.enabled = self._has_valid_filepath()
        op = export_row.operator("mastro_schedule.excel_export")
        op.node_name = self.name
        row = layout.row()
        row.template_list(
            "MASTRO_UL_schedule_export_sheets", "", self, "sheet_items", self, "active_sheet_index", rows=4,
        )
        col = row.column(align=True)
        op = col.operator("mastro_schedule.export_sheets_move", icon='TRIA_UP', text="")
        op.node_name = self.name
        op.direction = 'UP'
        op = col.operator("mastro_schedule.export_sheets_move", icon='TRIA_DOWN', text="")
        op.node_name = self.name
        op.direction = 'DOWN'

    def evaluate(self, inputs):
        # No real output of its own (this is the end of the chain,
        # like the Viewer) - nothing to compute or cache here.
        # export_sheets() below reads each linked Sheet straight from
        # execution.py's own evaluation cache (get_node_table) instead
        # of anything cached on this node, the same way
        # _sync_sheet_items already does for sheet_items' own labels -
        # more robust than stashing inputs[0] on a plain Python
        # attribute, which wouldn't survive Blender recreating this
        # node's own Python instance (e.g. across an undo).
        return []

    def _current_sheets_by_key(self):
        """Rebuilds the link_key -> Sheet value mapping straight from
        the evaluation cache (get_node_table). The KEY still comes
        from resolve_origin_node, matching whatever sheet_items itself
        was keyed by - but the VALUE comes from resolve_through_reroutes
        instead (the actual node directly feeding this link, e.g. Move
        Sheet/Sheet Background, not walked through the way the key's
        own origin is) - confirmed live as a real bug otherwise: using
        the origin for the value too made Export Excel skip every
        transparent operator's own work entirely (a Sheet moved by
        Move Sheet, or colored by Sheet Background, got exported as
        if neither had ever run - only the bare Sheet Primitive's own
        original output). Same identity/value split Join Tables/Place
        in Sheet's own _sync_table_items already makes for their
        labels - see nodes_table_join.py's own comment for the same
        reasoning applied there."""
        socket = self.inputs["Sheet"]
        sheets_by_key = {}
        for link_position, link in enumerate(socket.links):
            if link.is_muted:
                continue
            origin_node, origin_socket = resolve_origin_node(link)
            if origin_node is None:
                continue
            try:
                origin_output_index = list(origin_node.outputs).index(origin_socket)
            except ValueError:
                continue
            value_node, value_socket = resolve_through_reroutes(link)
            if value_node is None:
                continue
            try:
                value_output_index = list(value_node.outputs).index(value_socket)
            except ValueError:
                continue
            sheet = get_node_table(self.id_data.name, value_node.name)
            if sheet is None:
                continue
            key = _link_key(origin_node, origin_output_index, link_position)
            sheets_by_key[key] = sheet[value_output_index]
        return sheets_by_key

    def export_sheets(self):
        """Writes every entry in sheet_items (in that list's own
        order) into one .xlsx workbook at self.filepath, via
        excel_export_shared.write_sheet/save_workbook_atomically.
        Called only from MASTRO_OT_Schedule_Excel_Export - the manual
        Export button is the only way this ever runs (see this
        class's own module comment for why an auto_export option was
        tried and removed)."""
        import os
        import bpy
        from openpyxl import Workbook, load_workbook
        from .excel_export_shared import write_sheet, save_workbook_atomically

        sheets_by_key = self._current_sheets_by_key()
        path = bpy.path.abspath(self.filepath)
        # Force a real .xlsx extension - confirmed live as a real
        # surprise otherwise: a filepath the user typed without one
        # (or with the wrong one) still got written to verbatim,
        # leaving a file some file managers/OSes don't recognize as an
        # Excel workbook at all (a .xlsx is technically just a ZIP
        # archive of XML - without the right extension some tools show
        # it as a bare .zip instead).
        if not path.lower().endswith(".xlsx"):
            path = os.path.splitext(path)[0] + ".xlsx"
        # UPDATE mode only makes sense against the file's own existing
        # content - load_workbook() opens it as-is if it already
        # exists, so write_sheet's own UPDATE branch (which expects to
        # find a same-named sheet already there) has something real
        # to update. A fresh Workbook() is used otherwise (first
        # export, or REPLACE mode rebuilding from nothing anyway).
        if os.path.exists(path):
            workbook = load_workbook(path)
            default_sheet = None
        else:
            workbook = Workbook()
            # openpyxl's own Workbook() starts with one default sheet
            # ("Sheet") - removed once at least one real sheet has
            # been written, so an export with linked Sheets never
            # leaves that stray empty default sheet behind; if nothing
            # was linked at all, it's left in place rather than saving
            # a workbook with zero sheets (which openpyxl itself
            # refuses to do).
            default_sheet = workbook.active
        wrote_any = False
        for index, item in enumerate(self.sheet_items):
            sheet_data = sheets_by_key.get(item.link_key)
            if sheet_data is None:
                continue
            # item.label - this entry's own resolved name, already
            # following table_or_sheet_name set upstream (Table to
            # Sheet/Join Tables/Join Sheets), see _origin_label/
            # _sync_sheet_items above - not a separate editable field
            # here anymore (the user's own explicit removal: renaming
            # happens once, upstream, not a second time on this list).
            # "Sheet {index+1}" (1-based, this entry's own position in
            # the UIList) when even that's blank - NOT openpyxl's own
            # create_sheet() default (which always invents a brand new
            # "Sheet"/"Sheet1"/"Sheet2"/... name every single call,
            # confirmed live as a real bug: every export with a blank
            # name kept ADDING a new sheet instead of ever overwriting
            # the one from the previous export, accumulating forever).
            # A name tied to this entry's own stable position means the
            # same entry always writes to the same sheet, blank name or
            # not.
            sheet_name = item.label or f"Sheet {index + 1}"
            write_sheet(workbook, sheet_data, sheet_name, item.update_mode, item.start_cell)
            wrote_any = True
        if wrote_any and default_sheet is not None:
            workbook.remove(default_sheet)

        os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
        save_workbook_atomically(workbook, path)

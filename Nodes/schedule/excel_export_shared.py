"""Shared conversion/writing logic for Export Excel
(nodes_excel_export.py) - kept separate from the node itself so the
node module stays focused on the Blender-side (sockets, properties,
UI), mirroring table_text_edit_shared.py's own separation for the
Cells nodes.

No formulas anywhere - the user's own explicit call: this only ever
writes static values/style (text, fill color, font color, merged
cells), never an Excel formula. Color conversion is a direct linear
RGB float (0.0-1.0, Blender's own convention) -> int 0-255 -> hex
mapping, no gamma/color-management correction - the user's own
explicit call, the same color value shown in the Viewer becomes the
same hex in Excel.
"""

import re

_CELL_REF_RE = re.compile(r"^([A-Za-z]+)([0-9]+)$")


def parse_cell_ref(ref):
    """"B3" -> (row=2, col=1), 0-indexed (row 0 = Excel row 1, col 0 =
    column A) to match this tree's own Sheet/Table row/col convention
    elsewhere (merges, _link_key, ...). Returns (0, 0) for anything
    that doesn't parse as a plain column-letters + row-number
    reference - the same "fall back to the sane default rather than
    raise" rule every other user-typed-text field in this tree follows
    (e.g. Table Primitive's own Title pattern parsing)."""
    match = _CELL_REF_RE.match((ref or "").strip())
    if not match:
        return 0, 0
    letters, digits = match.groups()
    col = 0
    for ch in letters.upper():
        col = col * 26 + (ord(ch) - ord('A') + 1)
    return int(digits) - 1, col - 1


def rgb_to_hex(rgb):
    """(r, g, b) floats in 0.0-1.0 -> "RRGGBB" hex string, openpyxl's
    own color format - direct linear mapping, no gamma correction (see
    this module's own docstring for why)."""
    r, g, b = rgb
    return "".join(f"{max(0, min(255, round(c * 255))):02X}" for c in (r, g, b))


def write_sheet(workbook, sheet_data, sheet_name, update_mode, start_cell):
    """Writes one Sheet value (see sockets.py:MaStroScheduleSheetSocket
    for its {"columns": [{"cells": [...]}], "merges": [...]} shape)
    into `workbook` (an openpyxl Workbook), under `sheet_name` (or
    Excel's own default name if falsy).

    update_mode=False (REPLACE): any existing sheet with this name is
    removed and recreated from scratch.
    update_mode=True (UPDATE): an existing sheet with this name is
    kept, but every cell at row>=start row AND column>=start column
    (both from parse_cell_ref(start_cell)) is cleared first - the
    user's own explicit design, so content the user added directly in
    Excel above/to the left of the write area survives every
    re-export."""
    from openpyxl.styles import PatternFill, Font, Border, Side

    if sheet_name and sheet_name in workbook.sheetnames:
        if not update_mode:
            del workbook[sheet_name]
            ws = workbook.create_sheet(sheet_name)
        else:
            ws = workbook[sheet_name]
            start_row, start_col = parse_cell_ref(start_cell)
            # Existing merges that fall (even partially) inside the
            # rectangle about to be cleared must be un-merged FIRST -
            # openpyxl raises "MergedCell is read-only" the moment a
            # later write touches any cell still covered by a merge.
            # min_row/min_col are openpyxl's own 1-indexed coordinates;
            # start_row/start_col from parse_cell_ref are 0-indexed
            # (this tree's own convention elsewhere), hence the +1.
            for merge_range in list(ws.merged_cells.ranges):
                if merge_range.min_row >= start_row + 1 or merge_range.min_col >= start_col + 1:
                    ws.unmerge_cells(str(merge_range))
            for row in ws.iter_rows(min_row=start_row + 1, min_col=start_col + 1):
                for cell in row:
                    # "" not None - same fix as below, for the same
                    # reason (a None-valued, unstyled cell gets
                    # silently dropped from the saved file).
                    cell.value = ""
                    cell.fill = PatternFill(fill_type=None)
                    cell.font = Font()
    elif sheet_name:
        ws = workbook.create_sheet(sheet_name)
    else:
        ws = workbook.create_sheet()
    # Explicit True, not left at openpyxl's own default of None/absent
    # - confirmed live as a real Gnumeric-specific rendering gap: a
    # freshly created worksheet's <sheetView> has no showGridLines
    # attribute at all unless this is set explicitly, and Gnumeric (at
    # least) renders that absence as "no gridlines" rather than the
    # OOXML spec's own "absent means true" default - explicit beats
    # implicit here regardless of which reader is being strict about it.
    ws.sheet_view.showGridLines = True

    start_row, start_col = parse_cell_ref(start_cell) if update_mode else (0, 0)

    def write_cell(row_index, col_index, cell):
        # "" not None - confirmed live as a real bug otherwise:
        # openpyxl silently drops a cell from the saved file entirely
        # if its value is None AND it has no style of its own (a cell
        # with text but no fill/font survives a save/reload roundtrip
        # with max_row=1 instead of the real row count - confirmed by
        # writing then reloading a small test file). Every cell from a
        # Sheet always gets a real string value this way, so it's
        # always actually written, regardless of whether it ends up
        # with a fill/font too.
        ws_cell = ws.cell(
            row=start_row + row_index + 1, column=start_col + col_index + 1,
            value=cell.get("text") or "",
        )
        bg = cell.get("bg")
        if bg is not None:
            ws_cell.fill = PatternFill(start_color=rgb_to_hex(bg), end_color=rgb_to_hex(bg), fill_type="solid")
        text_color = cell.get("text_color")
        if text_color is not None:
            ws_cell.font = Font(color=rgb_to_hex(text_color))
        # Set by Sheet Grid (nodes_sheet_grid.py) - a real border drawn
        # on every side, the only way a grid stays visible over a
        # filled cell (showGridLines, set unconditionally above, is
        # masked by any cell's own background fill - confirmed live as
        # the actual reason gridlines disappeared behind colored
        # cells, both in Excel and Gnumeric, not a bug in either
        # reader).
        border = cell.get("border")
        if border is not None:
            side = Side(style=border.get("style", "thin"), color=rgb_to_hex(border.get("color", (0, 0, 0))))
            ws_cell.border = Border(left=side, right=side, top=side, bottom=side)

    # Read through the SAME header/rows shape the Viewer's own
    # _evaluate_table reads (sheet_to_table_shape, sheet_shared.py) -
    # one shared interpretation of "how to read a Sheet" used by both,
    # rather than this and the Viewer's own reading silently drifting
    # apart from each other over time. header is just row 0 here -
    # Sheet has no real header/row distinction (see
    # MaStroScheduleSheetSocket's own docstring), this is purely a
    # shape the conversion produces for code (this function, and the
    # Viewer's own _evaluate_table) that already knows how to read it.
    from .sheet_shared import sheet_to_table_shape
    table_shaped = sheet_to_table_shape(sheet_data)
    for col_index, column in enumerate(table_shaped.get("columns", [])):
        write_cell(0, col_index, column.get("header", {"text": "", "bg": None}))
        for row_index, cell in enumerate(column.get("rows", []), start=1):
            write_cell(row_index, col_index, cell)

    for merge in sheet_data.get("merges", []):
        ws.merge_cells(
            start_row=start_row + merge.get("start_row", 0) + 1,
            start_column=start_col + merge.get("start_col", 0) + 1,
            end_row=start_row + merge.get("end_row", 0) + 1,
            end_column=start_col + merge.get("end_col", 0) + 1,
        )
        text = merge.get("text")
        if text:
            ws.cell(row=start_row + merge.get("start_row", 0) + 1,
                     column=start_col + merge.get("start_col", 0) + 1, value=text)


def save_workbook_atomically(workbook, path):
    """Writes `workbook` to a temp file in the same directory as `path`,
    then os.replace()s it into place - the user's own explicit call,
    so a crash/kill mid-save never leaves a half-written .xlsx sitting
    at the real export path (os.replace is atomic on the same
    filesystem, which a same-directory temp file guarantees)."""
    import os
    import tempfile

    directory = os.path.dirname(os.path.abspath(path)) or "."
    fd, temp_path = tempfile.mkstemp(suffix=".xlsx", dir=directory)
    os.close(fd)
    try:
        workbook.save(temp_path)
        os.replace(temp_path, path)
    except Exception:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        raise
